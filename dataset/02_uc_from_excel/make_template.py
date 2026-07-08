#!/usr/bin/env python3
"""All-Japan-Grid — Excel で発電機を設定してユニットコミットメント(UC)を回すための
テンプレート .xlsx を生成する。

実在フリート (配布物 ``dist/matpower_national/<island>_genname.csv``) を初期値に、
UC に必要な運用パラメータ (限界費用・起動費・最小 up/down 時間・最小出力) を
``config/uc_config.yaml`` の既定値から埋めて、編集可能な 3 シートの Excel を作ります。

生成される ``generators_template.xlsx`` の 3 シート:
  - generators : 発電機 1 台 = 1 行 (id / name / fuel / bus / Pmax / Pmin / 限界費用 / …)
  - demand     : 24 時間の需要 (hour / demand_MW)
  - readme     : 各列の意味・単位・出典・編集方法

使い方::

    python make_template.py                     # okinawa (沖縄・最小フリート)
    python make_template.py --island okinawa --peak-mw 2000 --out generators_template.xlsx

⚠ 値の出所 (誠実性):
  - 発電機の名前 / 燃料 / Pmax は OSM 由来の建造モデル (実在フリート) です。
  - 限界費用・起動費・最小 up/down・最小出力 (Pmin) は **例題用の一般的な既定値**で、
    特定発電所の実測値ではありません (config/uc_config.yaml の typical estimates)。
    実データがある場合は Excel 上で書き換えて使ってください。
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parents[2]

# 燃料名の正規化 (建造モデルの表記 -> config/uc_config.yaml のキー)
FUEL_NORMALISE = {"gas": "lng"}
# 火力 (可制御) とみなす燃料。起動費・最小出力・最小 up/down を持たせる。
THERMAL_FUELS = {"coal", "lng", "oil", "nuclear", "biomass", "mixed"}


def _load_uc_defaults() -> dict:
    cfg = yaml.safe_load((ROOT / "config" / "uc_config.yaml").read_text(encoding="utf-8"))
    return cfg.get("defaults", {})


def _load_demand_shape(island: str, peak_mw: float | None) -> tuple[list[float], float]:
    """fy2023 シナリオの 24h 需要形状 × ピークを返す。"""
    scen = yaml.safe_load(
        (ROOT / "config" / "uc_scenarios" / "fy2023.yaml").read_text(encoding="utf-8")
    )
    shape = scen["demand"]["shape_24h"]
    if peak_mw is None:
        peaks = scen["demand"]["regional_peak_mw"]
        if island not in peaks:
            sys.exit(f"[ERROR] fy2023 に {island} のピークがありません。--peak-mw で指定してください。")
        peak_mw = float(peaks[island])
    return [round(peak_mw * s, 1) for s in shape], peak_mw


def _load_fleet(island: str) -> list[dict]:
    """配布物 <island>_genname.csv から可制御・RE 発電機の行を読む (slack 除外)。"""
    path = ROOT / "dist" / "matpower_national" / f"{island}_genname.csv"
    if not path.exists():
        sys.exit(f"[ERROR] フリート台帳が見つかりません: {path}")
    rows = list(csv.DictReader(path.open(encoding="utf-8")))
    return [r for r in rows if r.get("kind") == "gen"]


def build_rows(island: str) -> list[dict]:
    defaults = _load_uc_defaults()
    fuel_cost = defaults.get("fuel_cost_per_mwh", {})
    su = float(defaults.get("startup_cost", 5000))
    sd = float(defaults.get("shutdown_cost", 2000))
    mup = int(defaults.get("min_up_time_h", 4))
    mdn = int(defaults.get("min_down_time_h", 4))
    labor = float(defaults.get("labor_cost_per_h", 1000))

    out = []
    for i, r in enumerate(_load_fleet(island), start=1):
        raw_fuel = (r.get("fuel") or "unknown").strip().lower()
        fuel = FUEL_NORMALISE.get(raw_fuel, raw_fuel)
        pmax = round(float(r["PMAX"]), 1)
        is_thermal = fuel in THERMAL_FUELS
        out.append({
            "id": f"{island}_g{i:03d}",
            "name": r.get("name", ""),
            "fuel": fuel,
            "bus": int(float(r["GEN_BUS"])),
            "Pmax_MW": pmax,
            # 例題用の技術的最小出力の仮定 (火力=定格の30%, RE=0)。実測ではない。
            "Pmin_MW": round(pmax * 0.30, 1) if is_thermal else 0.0,
            "marginal_cost_JPY_per_MWh": float(fuel_cost.get(fuel, fuel_cost.get("unknown", 5000))),
            "startup_cost_JPY": su if is_thermal else 0.0,
            "shutdown_cost_JPY": sd if is_thermal else 0.0,
            "min_up_h": mup if is_thermal else 1,
            "min_down_h": mdn if is_thermal else 1,
            "no_load_cost_JPY_per_h": labor if is_thermal else 0.0,
            "init_on": 0,
        })
    return out


GEN_COLS = [
    ("id", "発電機ID (一意)"),
    ("name", "発電所名 (実在フリート・建造モデル由来)"),
    ("fuel", "燃料 (coal/lng/oil/nuclear/hydro/wind/solar/... )"),
    ("bus", "接続バス番号 (MATPOWER 配布ケースのバス)"),
    ("Pmax_MW", "定格出力 [MW] (Pmax・建造モデル由来)"),
    ("Pmin_MW", "最小出力 [MW] (例題仮定: 火力=定格30%・要編集)"),
    ("marginal_cost_JPY_per_MWh", "限界費用 [円/MWh] (燃料別の一般既定・要編集)"),
    ("startup_cost_JPY", "起動費 [円/回] (一般既定・要編集)"),
    ("shutdown_cost_JPY", "停止費 [円/回] (一般既定)"),
    ("min_up_h", "最小連続運転時間 [h]"),
    ("min_down_h", "最小連続停止時間 [h]"),
    ("no_load_cost_JPY_per_h", "無負荷固定費 [円/h] (起動中)"),
    ("init_on", "初期状態 (1=運転中 / 0=停止)"),
]

README_LINES = [
    ("All-Japan-Grid — Excel で発電機設定 → UC (ユニットコミットメント)", ""),
    ("", ""),
    ("このブックを編集して run_uc.py に渡すと、24 時間の最小コスト起動停止計画を解きます。", ""),
    ("  python run_uc.py --xlsx generators_template.xlsx", ""),
    ("", ""),
    ("【 generators シート 】発電機 1 台 = 1 行。列の意味:", ""),
] + [(f"  {c}", d) for c, d in GEN_COLS] + [
    ("", ""),
    ("【 demand シート 】24 時間の需要。hour=0..23, demand_MW=各時刻の需要[MW]。", ""),
    ("  発電機の Pmax 合計より需要が大きいと解が存在しません (infeasible)。", ""),
    ("", ""),
    ("【 値の出所・誠実性 】", ""),
    ("  発電所名/燃料/Pmax = OSM 由来の建造モデル (実在フリート)。", ""),
    ("  限界費用/起動費/最小up-down/Pmin = 例題用の一般既定 (config/uc_config.yaml)。", ""),
    ("  特定発電所の実測値ではありません。実データがあれば書き換えてください。", ""),
    ("  燃料コスト出典: 国内火力の典型推定値 (要較正)。", ""),
]


def write_xlsx(rows: list[dict], demand: list[float], out_path: Path, island: str, peak_mw: float) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2E5A88")

    # --- generators シート ---
    ws = wb.active
    ws.title = "generators"
    cols = [c for c, _ in GEN_COLS]
    ws.append(cols)
    for j in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=j)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r[c] for c in cols])
    ws.freeze_panes = "A2"
    for j, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(12, min(34, len(c) + 4))

    # --- demand シート ---
    wd = wb.create_sheet("demand")
    wd.append(["hour", "demand_MW"])
    for j in (1, 2):
        wd.cell(row=1, column=j).font = head_font
        wd.cell(row=1, column=j).fill = head_fill
    for h, d in enumerate(demand):
        wd.append([h, d])
    wd.freeze_panes = "A2"
    wd.column_dimensions["A"].width = 8
    wd.column_dimensions["B"].width = 14

    # --- readme シート ---
    wr = wb.create_sheet("readme")
    for term, desc in README_LINES:
        wr.append([term, desc])
    wr.column_dimensions["A"].width = 40
    wr.column_dimensions["B"].width = 60
    wr.cell(row=1, column=1).font = Font(bold=True, size=13)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--island", default="okinawa", help="対象島/地域 (既定: okinawa)")
    ap.add_argument("--peak-mw", type=float, default=None, help="ピーク需要 [MW] (既定: fy2023 の値)")
    ap.add_argument("--out", default=str(Path(__file__).parent / "generators_template.xlsx"),
                    help="出力 .xlsx パス")
    args = ap.parse_args()

    rows = build_rows(args.island)
    demand, peak = _load_demand_shape(args.island, args.peak_mw)
    out = Path(args.out)
    write_xlsx(rows, demand, out, args.island, peak)

    disp = sum(r["Pmax_MW"] for r in rows if r["fuel"] in THERMAL_FUELS)
    print(f"wrote {out}")
    print(f"  {len(rows)} generators ({args.island}), thermal Pmax={disp:.0f} MW, peak demand={peak:.0f} MW")
    if peak > disp:
        print("  ⚠ ピーク需要が火力容量を超えています。RE 込みでも infeasible の恐れ。--peak-mw を下げてください。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
