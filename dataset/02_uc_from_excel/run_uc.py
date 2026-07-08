#!/usr/bin/env python3
"""All-Japan-Grid — Excel の発電機設定を読み込んでユニットコミットメント(UC)を解く。

``make_template.py`` が生成した (または手で編集した) ``generators_template.xlsx`` を
入力に、24 時間の最小コスト起動停止計画 (どの発電機を・いつ・どれだけ動かすか) を
MILP で解き、結果を Excel と PNG グラフで出力します。

使い方::

    python run_uc.py                                   # generators_template.xlsx
    python run_uc.py --xlsx my_generators.xlsx --out result.xlsx

出力:
  - <out>.xlsx : dispatch シート (発電機×時刻の出力) + summary シート (コスト内訳)
  - <out>.png  : 燃料別の積み上げ発電量 + 需要曲線

依存: openpyxl, pulp, matplotlib, および本リポジトリの src/ (UC ソルバ)
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from src.model.generator import Generator  # noqa: E402
from src.uc.models import DemandProfile, TimeHorizon, UCParameters  # noqa: E402
from src.uc.solver import solve_uc  # noqa: E402

# 燃料 -> 表示色 (積み上げグラフ)。merit order の直感に沿った配色。
FUEL_COLOR = {
    "nuclear": "#7b3ff2", "coal": "#4d4d4d", "lng": "#3b82c4", "oil": "#c0504d",
    "hydro": "#2e86c1", "pumped_hydro": "#5dade2", "geothermal": "#af7ac5",
    "biomass": "#8b5e3c", "wind": "#27ae60", "solar": "#f1c40f",
    "mixed": "#95a5a6", "unknown": "#bdc3c7",
}


def load_inputs(xlsx: Path):
    from openpyxl import load_workbook

    if not xlsx.exists():
        sys.exit(f"[ERROR] 入力が見つかりません: {xlsx}\n        先に make_template.py を実行してください。")
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["generators"]
    header = [c.value for c in ws[1]]
    gens: list[Generator] = []
    init: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        if not d.get("id"):
            continue
        g = Generator(
            id=str(d["id"]),
            name=str(d.get("name") or d["id"]),
            capacity_mw=float(d["Pmax_MW"]),
            fuel_type=str(d.get("fuel") or "unknown"),
            connected_bus_id=str(d.get("bus", "")),
            p_min_mw=float(d.get("Pmin_MW") or 0.0),
            fuel_cost_per_mwh=float(d.get("marginal_cost_JPY_per_MWh") or 0.0),
            startup_cost=float(d.get("startup_cost_JPY") or 0.0),
            shutdown_cost=float(d.get("shutdown_cost_JPY") or 0.0),
            min_up_time_h=int(d.get("min_up_h") or 1),
            min_down_time_h=int(d.get("min_down_h") or 1),
            no_load_cost=float(d.get("no_load_cost_JPY_per_h") or 0.0),
        )
        gens.append(g)
        init[g.id] = int(d.get("init_on") or 0)
    wd = wb["demand"]
    demand = [float(r[1]) for r in wd.iter_rows(min_row=2, values_only=True) if r[1] is not None]
    return gens, demand, init


def write_result_xlsx(gens, result, demand, out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    name_by_id = {g.id: g.name for g in gens}
    fuel_by_id = {g.id: g.fuel_type_enum.value for g in gens}
    n = len(demand)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2E5A88")

    wb = Workbook()
    # dispatch シート: 発電機×時刻の出力 [MW]
    ws = wb.active
    ws.title = "dispatch"
    ws.append(["id", "name", "fuel"] + [f"h{h}" for h in range(n)])
    for j in range(1, n + 4):
        ws.cell(row=1, column=j).font = head_font
        ws.cell(row=1, column=j).fill = head_fill
    sched_by_id = {s.generator_id: s for s in result.schedules}
    for g in gens:
        s = sched_by_id.get(g.id)
        power = (s.power_output_mw if s else [0.0] * n)
        ws.append([g.id, g.name, fuel_by_id[g.id]] + [round(p, 1) for p in power])
    ws.append(["DEMAND", "(需要)", ""] + [round(d, 1) for d in demand])
    ws.freeze_panes = "D2"

    # summary シート: ステータス + 発電機別のエネルギー・コスト
    wsum = wb.create_sheet("summary")
    wsum.append(["status", result.status])
    wsum.append(["total_cost_JPY", round(result.total_cost, 1)])
    wsum.append(["solve_time_s", round(result.solve_time_s, 3)])
    wsum.append(["num_generators", result.num_generators])
    wsum.append([])
    wsum.append(["id", "name", "fuel", "energy_MWh", "startups", "fuel_cost", "startup_cost", "total_cost"])
    hdr = wsum.max_row
    for j in range(1, 9):
        wsum.cell(row=hdr, column=j).font = head_font
        wsum.cell(row=hdr, column=j).fill = head_fill
    for s in result.schedules:
        wsum.append([
            s.generator_id, name_by_id.get(s.generator_id, ""), fuel_by_id.get(s.generator_id, ""),
            round(s.total_energy_mwh, 1), s.num_startups,
            round(s.fuel_cost, 1), round(s.startup_cost, 1), round(s.total_cost, 1),
        ])
    for col, w in (("A", 16), ("B", 28), ("C", 10), ("D", 12), ("E", 9), ("F", 12), ("G", 12), ("H", 12)):
        wsum.column_dimensions[col].width = w

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def write_plot(gens, result, demand, out_png: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    n = len(demand)
    fuel_by_id = {g.id: g.fuel_type_enum.value for g in gens}
    # 燃料ごとに出力を合算
    by_fuel: dict[str, np.ndarray] = {}
    for s in result.schedules:
        f = fuel_by_id.get(s.generator_id, "unknown")
        arr = np.array(s.power_output_mw[:n] + [0.0] * (n - len(s.power_output_mw)))
        by_fuel[f] = by_fuel.get(f, np.zeros(n)) + arr

    # merit order (安い順) に積む: 表示は下から安い燃料
    order = ["nuclear", "hydro", "geothermal", "wind", "solar", "biomass",
             "coal", "lng", "oil", "pumped_hydro", "mixed", "unknown"]
    fuels = [f for f in order if f in by_fuel] + [f for f in by_fuel if f not in order]

    fig, ax = plt.subplots(figsize=(11, 6))
    hours = np.arange(n)
    stacks = [by_fuel[f] for f in fuels]
    colors = [FUEL_COLOR.get(f, "#bdc3c7") for f in fuels]
    ax.stackplot(hours, *stacks, labels=fuels, colors=colors, alpha=0.9)
    ax.plot(hours, demand, "k--", lw=2, label="demand")
    ax.set_xlabel("hour")
    ax.set_ylabel("power [MW]")
    ax.set_title(f"Unit Commitment dispatch  (status={result.status}, "
                 f"total cost={result.total_cost:,.0f} JPY)")
    ax.set_xlim(0, n - 1)
    ax.legend(loc="upper left", ncol=2, fontsize=8)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=130)
    plt.close(fig)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    here = Path(__file__).parent
    ap.add_argument("--xlsx", default=str(here / "generators_template.xlsx"), help="入力 Excel")
    ap.add_argument("--out", default=str(here / "uc_result.xlsx"), help="出力 Excel (.png も同名で生成)")
    ap.add_argument("--reserve", type=float, default=0.05, help="予備率 (既定 0.05 = 5%%)")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    gens, demand, init = load_inputs(xlsx)
    print(f"loaded {len(gens)} generators, {len(demand)} h demand (peak={max(demand):.0f} MW) from {xlsx.name}")

    params = UCParameters(
        generators=gens,
        demand=DemandProfile(demands=demand),
        time_horizon=TimeHorizon(num_periods=len(demand), period_duration_h=1.0),
        reserve_margin=args.reserve,
        initial_commitment=init,
        mip_gap=0.01,
        solver_time_limit_s=120,
    )
    result = solve_uc(params)
    print(f"UC status: {result.status}  total_cost={result.total_cost:,.0f} JPY  "
          f"solve={result.solve_time_s:.2f}s")
    if result.warnings:
        for w in result.warnings[:5]:
            print(f"  warning: {w}")
    if not result.is_optimal and result.status != "Optimal":
        print(f"  (status={result.status} — 需要が容量を超えていないか demand シートを確認してください)")

    out = Path(args.out)
    write_result_xlsx(gens, result, demand, out)
    print(f"wrote {out}")
    png = out.with_suffix(".png")
    try:
        write_plot(gens, result, demand, png)
        print(f"wrote {png}")
    except ImportError:
        print("  (matplotlib が無いため png はスキップ — `pip install matplotlib` で出力できます)")
    return 0 if result.is_optimal else 1


if __name__ == "__main__":
    raise SystemExit(main())
